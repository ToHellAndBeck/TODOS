from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import openpyxl
from openpyxl import Workbook

def read_numbers_from_excel(file_path, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    numbers = []
    for cell in sheet[column]:
        if cell.value is not None:
            numbers.append(str(cell.value))
    return numbers

def write_to_excel(file_path, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, (text1, text2) in enumerate(data, start=1):
        sheet.cell(row=i, column=1).value = text1
        sheet.cell(row=i, column=2).value = text2
    workbook.save(file_path)

def wait_for_element_to_be_clickable(driver, xpath):
    try:
        button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        return button
    except Exception as e:
        print(xpath, e)
 
def click_button(driver, xpath):
    try:
        button = wait_for_element_to_be_clickable(driver, xpath)
        if button:
            button.click()
    except Exception as e:
        print(xpath, e)
 
def input_text(driver, element_id, text):
    try:
        # Find the email input field by its ID
        email_input = driver.find_element(By.ID, element_id)
        # Clear the input field if needed
        email_input.clear()
        # Input the email address
        email_input.send_keys(text)
        email_input.send_keys(Keys.ENTER)
        print(f"Entered text: {text}")
    except Exception as e:
        print(f"An error occurred: {e}")
        
def read_excel_sheet(excel_path):
    print(f"Reading Excel file: {excel_path}")  # Diagnostic print
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            first_name = row[0]
            last_name = row[1]
            email = row[2]
            data.append((first_name, last_name, email))
        print("Data read from Excel:", data)
        return data
    except Exception as e:
        print(f"Error reading Excel file: {e}")

# Function to save the data to an Excel file
def save_data_to_excel(data, filename):
    workbook = Workbook()
    sheet = workbook.active
    headers = ['Username', 'Full Name', 'Email', 'User Status', 'User Type', 'Customer', 'City', 'State']
    sheet.append(headers)

    for row_data in data:
        sheet.append(row_data)

    workbook.save(filename)

    
def enter_and_process_names(driver, names):
    for first_name, last_name, email in names:
        try:
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_txtFirstName")))
            first_name_field = driver.find_element(By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_txtFirstName")
            first_name_field.clear()
            first_name_field.send_keys(first_name)
            print(f"Entering first name: {first_name}")  # Debug print
            
            last_name_field = driver.find_element(By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_txtLastName")
            last_name_field.clear()
            last_name_field.send_keys(last_name)
            print(f"Entering last name: {last_name}")  # Debug print
            last_name_field.send_keys(Keys.ENTER)

            WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_grdUsers"))
            )
            print(f"Entered and submitted: {first_name} {last_name}")
            
        except Exception as e:
            print(f"Failed to enter name {first_name} {last_name} due to: {e}")
        time.sleep(3)
        
        try:
            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_rgrdUsers_ctl00"))
            )
            print("Table is present.")

            rows = driver.find_elements(By.CSS_SELECTOR, "#ctl00_cphMain_rcUserAdmin_UserAdminControl1_rgrdUsers_ctl00 tbody tr")
            print(f"Found {len(rows)} rows in the table.")
            active_user_data = []
            for row in rows:
                columns = row.find_elements(By.TAG_NAME, "td")
                try:
                    status = columns[4].text.strip()  # Assuming the status is in the 5th column
                    if status == "Active":
                        # Capture the relevant data from each column
                        user_data = [col.text.strip() for col in columns[1:9]]  # Adjust indices if needed
                        print(f"Active user found: {user_data}")
                        active_user_data.append(user_data)

                        # Find the pencil icon in the first cell and click it
                        edit_icon = columns[0].find_element(By.CSS_SELECTOR, "a[id*='hypEdit']")
                        edit_icon.click()
                        print(f"Clicked the pencil icon for: {user_data[0]}")  # Assuming first item is the username
                        time.sleep(3)
                except (IndexError, NoSuchElementException) as e:
                    print(f"Skipping to security tab due to: {e}")
                    break  # Exit the loop if we can't find the active status or pencil icon
        except Exception as e:
            print(f"No table found or error occurred: {e}")
        
        # Proceed to the security tab and external apps tab
        security_tab_xpath = "/html/body/form/div[7]/span/div/div/div/div[1]/div/ul/li[2]/a/span/span/span"
        click_button(driver, security_tab_xpath)
        print("Clicked Security Tab")
        time.sleep(3)
        email_xpath = "//div[@id='ctl00_ctl00_cphMain_rc_ProfileContent_ProfileMgmtSecurity1_rts']/div/ul/li[4]/a/span/span/span"
        click_button(driver, email_xpath)
        print("Clicked the External Apps tab")
        time.sleep(2)
        external_app_input_field_xpath = "//input[@id='ctl00_ctl00_cphMain_rc_ProfileContent_ProfileMgmtSecurity1_ProfileExternalApp_txtAdUsername']"
        external_app_email = driver.find_element(By.XPATH, external_app_input_field_xpath)
        external_app_email.clear()
        external_app_email.send_keys(email)
        print(f"Entering email: {email}")
        savebuttonXpath='//*[@id="ctl00_ctl00_cphMain_rc_ProfileContent_ProfileMgmtSecurity1_ProfileExternalApp_btnSave"]'
        click_button(driver, savebuttonXpath)
        print(f"Saved email, {email}") 
        time.sleep(3)
                
        # Return to the original URL before processing the next name
        driver.get(url)
        print("Returned to the original URL.")
        
 

        

# Initialize the Chrome driver
# chrome_options=Options()
# chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))#,options=chrome_options
url = 'https://trust.wachter.com/Admin/UserAdmin.aspx'
employee_dropdown_xpath = '//*[@id="ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input"]'
rcb_list_xpath = '/html/body/form/div[1]/div/div/ul'
email_element_id = 'i0116'
dropdown_input_id = 'ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input'
driver.get(url)
login_button_xpath = '//*[@id="slide-5-layer-8"]'
click_button(driver, login_button_xpath)
wachter_login_button_xpath = '//*[@id="ctl00_cphMain_ssoLoginControl_btnSingleSignOnLogin"]'
click_button(driver, wachter_login_button_xpath)
next_button_path = '//*[@id="idSIButton9"]'
wait_for_element_to_be_clickable(driver, next_button_path)
email_address = '****'
input_text(driver, email_element_id, email_address)
sign_in_button_xpath = '//*[@id="idSIButton9"]'
wait_for_element_to_be_clickable(driver, sign_in_button_xpath)
password_element_id = 'i0118'
password = '****'
time.sleep(1)
input_text(driver, password_element_id, password)
stay_signed_in_button_xpath = '//*[@id="idSIButton9"]'
click_button(driver, stay_signed_in_button_xpath)
time.sleep(3)
# Read names from an Excel sheet
excel_path = r"C:\Users\beckett.mcfarland\scripts\python\External_Linker\emails.xlsx"
names = read_excel_sheet(excel_path)

# Enter each name into the webpage and process active users
enter_and_process_names(driver, names)

# Close the driver
driver.quit()
