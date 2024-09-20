from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl

def read_numbers_from_excel(file_path, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    numbers = []
    for cell in sheet[column]:
        if cell.value is not None:
            numbers.append(str(cell.value))
    return numbers

def input_text(driver, element_id, text):
    try:
        email_input = driver.find_element(By.ID, element_id)
        email_input.clear()
        email_input.send_keys(text)
        email_input.send_keys(Keys.ENTER)
    except Exception as e:
        print(f"An error occurred: {e}")

def format_phone_number(number):
    # Ensure the number has 10 digits and is properly formatted
    cleaned_number = ''.join(filter(str.isdigit, number))  # Remove any non-digit characters
    if len(cleaned_number) == 10:
        return f"{cleaned_number[:3]}-{cleaned_number[3:6]}-{cleaned_number[6:]}"
    return number

def find_and_click_phone_number_link(driver, number):
    try:
        formatted_number = format_phone_number(number)
        print(f"Trying to find and click link with text: {formatted_number}")
        
        # Wait for the link containing the phone number to be visible
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.LINK_TEXT, formatted_number))
        )

        # Find the link with the phone number and click it
        link = driver.find_element(By.LINK_TEXT, formatted_number)
        link.click()
        return link.text
    except Exception as e:
        print(f"Error finding and clicking phone number link: {e}")
        return None

def get_text_from_class(driver, class_name):
    data = []
    try:
        elements = driver.find_elements(By.CLASS_NAME, class_name)
        for element in elements:
            data.append(element.text)
    except Exception as e:
        print(f"Error getting text from class {class_name}: {e}")
    return data

def write_to_excel(file_path, data, headers):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num).value = header

    for row_num, row_data in enumerate(data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num).value = value

    max_column = sheet.max_column
    columns_to_delete = []
    for col_num in range(1, max_column + 1):
        if sheet.cell(row=1, column=col_num).value == "":
            columns_to_delete.append(col_num)

    for col_num in reversed(columns_to_delete):
        sheet.delete_cols(col_num)

    workbook.save(file_path)

def wait_for_element_to_be_clickable(driver, xpath):
    try:
        button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        return button
    except Exception as e:
        print(xpath, e)

def click_button(driver, xpath):
    try:
        button = wait_for_element_to_be_clickable(driver, xpath)
        if button:
            button.click()
    except Exception as e:
        print(xpath, e)


# Initialize the Chrome driver
# chrome_options=Options()
# chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))#,options=chrome_options
url = 'https://linqservices.my.site.com/qonnect/s/'

def get_text_from_next_column(row_element, cell_index):
    try:
        next_column_xpath = f'./td[{cell_index + 2}]'  # Increment by 2 to get the next cell
        return row_element.find_element(By.XPATH, next_column_xpath).text
    except Exception as e:
        print(f"Error getting text from next column: {e}")
        return None
    
#dropdown_xpath = '//*[@id="ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input"]'
rcb_list_xpath = '/html/body/form/div[1]/div/div/ul'
email_element_id = '67:2;a'
driver.get(url)
time.sleep(3)
email_address = 'beckett.mcfarland@wachter.com'
input_text(driver, email_element_id, email_address)
time.sleep(2)
#sign_in_button_xpath = '//*[@id="idSIButton9"]'
#wait_for_element_to_be_clickable(driver, sign_in_button_xpath)
password_element_id = '80:2;a'
password = 'Ozymandias99!'
input_text(driver, password_element_id, password)
stay_signed_in_button_xpath = '//*[@id="idSIButton9"]'
click_button(driver, stay_signed_in_button_xpath)
time.sleep(5)
search_box_element_id = '178:0'


# Read numbers from the Excel file
excel_file_path = r'c:\Users\beckett.mcfarland\Documents\phone_numbers(input).xlsx'
numbers = read_numbers_from_excel(excel_file_path, 'A')

# IDs to copy text from
class_name_to_copy = 'uiOutputText'
headers = ["PHONE NUMBER", "NAME", "IMEI", "ICCID", "EMPLOYEE NUMBER", "DEPT CODE", "SUB CODE", "EMAIL"]

# Data to write to Excel
data_to_write = []

class_name_to_copy = 'uiOutputText'

# Data to write to Excel
data_to_write = []

for number in numbers:
    # Refresh the page or navigate back to the search page
    driver.get(f"https://linqservices.my.site.com/qonnect/s/global-search/{number}")
    time.sleep(4)

    # Click on the first phone number link
    phone_number = find_and_click_phone_number_link(driver, number)
    time.sleep(6)
    if phone_number:
        time.sleep(2)  # Wait for the page to load after clicking the link
        
        # Get text from specified class name
        extracted_texts = get_text_from_class(driver, class_name_to_copy)
        
        # Add the extracted texts directly as a row
        data_to_write.append(extracted_texts)
    else:
        print(f"No phone number link found for search number {number}")

# Write the collected data to an Excel file
output_excel_file_path = r'c:\Users\beckett.mcfarland\Documents\output_data.xlsx'
write_to_excel(output_excel_file_path, data_to_write)

# Close the driver
driver.quit()