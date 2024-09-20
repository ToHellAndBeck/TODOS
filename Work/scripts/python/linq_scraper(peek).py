from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl



def read_numbers_from_excel(file_path, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    numbers = []
    for cell in sheet[column]:
        if cell.value is not None:
            numbers.append(str(cell.value))
    return numbers

def input_text(driver, element_id, text):
    try:
        email_input = driver.find_element(By.ID, element_id)
        email_input.clear()
        email_input.send_keys(text)
        email_input.send_keys(Keys.ENTER)
    except Exception as e:
        print(f"An error occurred: {e}")

def extract_table_data(driver):
    try:
        # Wait for the second table's tbody element to be visible
        table_elements = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, '//tbody[@data-aura-rendered-by]'))
        )
        # The second table should be the second element in the list
        table_element = table_elements[1]
        
        # Extract table rows
        rows = table_element.find_elements(By.TAG_NAME, "tr")
        print(f"Found {len(rows)} rows in the table")
        
        extracted_data = []
        
        for row in rows:
            row_data = []
            cells = row.find_elements(By.TAG_NAME, "td")
            for cell in cells:
                # Extract text from each cell
                cell_text = cell.text.strip()
                # If the cell contains a link, extract the text from the link
                link = cell.find_elements(By.TAG_NAME, "a")
                if link:
                    cell_text = link[0].text.strip()
                row_data.append(cell_text)
                print(cell_text)  # Debugging: Print each cell's text
            extracted_data.append(row_data)
        
        return extracted_data
    except Exception as e:
        print(f"Error extracting table data: {e}")
        return []

def write_to_excel(file_path, data):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

    for row_data in data:
        sheet.append(row_data)

    workbook.save(file_path)

def click_button(driver, xpath):
    try:
        button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        button.click()
    except Exception as e:
        print(xpath, e)

# Initialize the Chrome driver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
url = 'https://linqservices.my.site.com/qonnect/s/'

# Login details
email_element_id = '67:2;a'
password_element_id = '80:2;a'
email_address = 'beckett.mcfarland@wachter.com'
password = 'Ozymandias99!'

driver.get(url)
time.sleep(3)
input_text(driver, email_element_id, email_address)
time.sleep(2)
input_text(driver, password_element_id, password)
stay_signed_in_button_xpath = '//*[@id="idSIButton9"]'
click_button(driver, stay_signed_in_button_xpath)
time.sleep(5)
search_box_element_id = '178:0'


# Read numbers from the Excel file
excel_file_path = r'c:\Users\beckett.mcfarland\Documents\phone_numbers(input).xlsx'
numbers = read_numbers_from_excel(excel_file_path, 'A')

# Output file path
output_excel_file_path = r'c:\Users\beckett.mcfarland\Documents\output_data.xlsx'

for number in numbers:
    # Navigate to the search page
    try:
        driver.get(f"https://linqservices.my.site.com/qonnect/s/global-search/{number}")
        time.sleep(4)

        # Extract data from the table
        table_data = extract_table_data(driver)
        if table_data:
            write_to_excel(output_excel_file_path, table_data)
        else:
            print(f"No table found for search number {number}")
    except Exception as e:
        print(f"Error navigating to search number {number}: {e}")

# Close the driver
driver.quit()